# Scrapes Switch game titles and prices from Gamestop and adds a margin of 10% to the price
# 	then saving it into an .xlsx file
# Written by Braeden Richards
# Created: March 7, 2018

from urllib.request import urlopen as uReq
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

url = 'https://www.gamestop.com/browse/games/nintendo-switch?nav=28-xu0,13ffff2418-1e8'
url_base = 'https://www.gamestop.com'

titles = []
prices = [] 
condition = [] # New/Used/Download

page_counter = 0
all_pages_done_flag = False # True when there are no more pages to scrape
while(not all_pages_done_flag):
	page_counter += 1
	print("Starting page #" + " : " + url + "\n")

	uClient = uReq(url)
	page_html = uClient.read()
	uClient.close()

	url = ''

	soup = BeautifulSoup(page_html, "html.parser")

	products = soup.findAll("div", {"class" : "product"})
	pricing = soup.findAll("div", {"class" : "purchase_info"})
 
	for product in products:
		titles.append(product.find("h3", {"class" : "ats-product-title"}).a.getText())

	for price in pricing:
		prices.append(price.p.getText())
		condition.append(price.h4.find("strong").getText())

	if (soup.find("div", {"class" : "pagination_controls"}).find("a", {"class" : "next_page"})) is None:
		all_pages_done_flag = True
	else:
		url = url_base + soup.find("div", {"class" : "pagination_controls"}).find("a", {"class" : "next_page"}).get("href")


# Open Excel workbook object
wb = Workbook()
ws = wb.active

# Headers
header_title = "Title"
header_condition = "Condition"
header_price = "Price"
header_markup = "Total"

# Excel Column Selection
cell_row = 1
cell_col_title = 'A'
cell_col_condition = 'B'
cell_col_price = 'C'
cell_col_markup = 'D'

# Inputs the titles of each column to the .xlsx file
ws[str(cell_col_title) + str(cell_row)] = header_title
ws[str(cell_col_title) + str(cell_row)].font = Font(bold=True,size=15,underline='single')
ws[str(cell_col_title) + str(cell_row)].alignment = Alignment(vertical='center', horizontal='center')

ws[str(cell_col_condition) + str(cell_row)] = header_condition
ws[str(cell_col_condition) + str(cell_row)].font = Font(bold=True,size=15,underline='single')
ws[str(cell_col_condition) + str(cell_row)].alignment = Alignment(vertical='center', horizontal='center')


ws[str(cell_col_price) + str(cell_row)] = header_price
ws[str(cell_col_price) + str(cell_row)].font = Font(bold=True,size=15,underline='single')
ws[str(cell_col_price) + str(cell_row)].alignment = Alignment(vertical='center', horizontal='center')

ws[str(cell_col_markup) + str(cell_row)] = header_markup
ws[str(cell_col_markup) + str(cell_row)].font = Font(bold=True,size=15,underline='single')
ws[str(cell_col_markup) + str(cell_row)].alignment = Alignment(vertical='center', horizontal='center')

# Inputting the data to the excel file
cell_row += 1
i = 0
for title in titles:
	stock_price = float(("$".join(prices[i].split("$", 2)[:2])).replace("$", ""))
	ws[str(cell_col_title) + str(cell_row)] = title
	ws[str(cell_col_title) + str(cell_row)].fill = PatternFill(fgColor = "41f4e8", fill_type = "solid")
	ws[str(cell_col_title) + str(cell_row)].font = Font(size=13)
	ws[str(cell_col_title) + str(cell_row)].alignment = Alignment(vertical='center', horizontal='left')

	ws[str(cell_col_condition) + str(cell_row)] = condition[i]
	ws[str(cell_col_condition) + str(cell_row)].fill = PatternFill(fgColor="41dff4", fill_type = "solid")
	ws[str(cell_col_condition) + str(cell_row)].font = Font(size=13)
	ws[str(cell_col_condition) + str(cell_row)].alignment = Alignment(vertical='center', horizontal='left')

	ws[str(cell_col_price) + str(cell_row)] = stock_price
	ws[str(cell_col_price) + str(cell_row)].fill = PatternFill(fgColor="f45941", fill_type = "solid")
	ws[str(cell_col_price) + str(cell_row)].font = Font(size=13)
	ws[str(cell_col_price) + str(cell_row)].alignment = Alignment(vertical='center', horizontal='center')
	
	ws[str(cell_col_markup) + str(cell_row)] = (stock_price * 0.8265)
	ws[str(cell_col_markup) + str(cell_row)].fill = PatternFill(fgColor="53f442", fill_type = "solid")
	ws[str(cell_col_markup) + str(cell_row)].font = Font(size=13)
	ws[str(cell_col_markup) + str(cell_row)].alignment = Alignment(vertical='center', horizontal='center')
	
	i += 1
	cell_row += 1

wb.save("switch_games.xlsx")
